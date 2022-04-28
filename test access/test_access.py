import re
from openpyxl import Workbook
import os

path="C:\\Users\\my laptop\\Desktop\\test_server_files"
dir_list = os.listdir(path)

book = Workbook()
for file in dir_list:
    ws1 = book.create_sheet(file)
    ws1.title = file
    ws1.cell(row=1, column=1).value = "Destination IP"
    ws1.cell(row=1, column=2).value = "State"
    count=2
    with open(path+"\\"+file) as f:
        for line in f:
            if re.search("^ComputerName",line):
                dst_ip=re.split("ComputerName.*:.",line)
                ws1.cell(row=count, column=1).value=dst_ip[1] 
                
            if re.search("^TcpTestSucceeded",line):
                stat=re.split("TcpTestSucceeded.*:.",line)
                ws1.cell(row=count, column=2).value=stat[1]
                count =count+1


    book.save("sample.xlsx")
